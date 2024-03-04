#!/usr/bin/env python
# coding: utf-8

# In[1]:


import face_recognition as fr
import cv2 as cv
from tkinter import Tk, Button, Label, Frame, filedialog, messagebox
import os
import openpyxl
from datetime import datetime

class FaceRecognitionGUI:
    def __init__(self, master):
        self.master = master
        master.title("Face Recognition")
        master.configure(background='white')

        self.label = Label(master, text="Smart Attendance Management System.", bg="blue", fg="black", font=("Times New roman", 36))
        self.label.pack(pady=5)

        self.label = Label(master, text="Click 'Choose Image' to select an image.",bg="pink", fg="black",font=("Arial", 16))
        self.label.pack(pady=5)

        self.choose_image_button = Button(master, text="Choose Image",bg="Red", fg="black", font=("Arial", 16), command=self.choose_image,padx=50 , pady=10)
        self.choose_image_button.pack(pady=10)

        self.attendance_button = Button(master, text="Take Attendance",bg="light green", fg="black",font=("Arial", 16), command=self.take_attendance, state='disabled',padx=50, pady=10)
        self.attendance_button.pack(pady=10)

        self.show_attendance_button = Button(master, text="Show Attendance",bg="light green", fg="black",font=("Arial", 16), command=self.show_attendance, state='disabled',padx=50, pady=10)
        self.show_attendance_button.pack(pady=10)

        self.exit_button = Button(master, text="Exit", command=self.master.destroy, bg='Red',font=("Arial",16), padx=50, pady=10)
        self.exit_button.pack(pady=10)

    def choose_image(self):
        Tk().withdraw()
        load_image = filedialog.askopenfilename()

        if load_image:
            self.target_image = fr.load_image_file(load_image)
            self.target_encoding = fr.face_encodings(self.target_image)
            self.label.configure(text="Image selected. Click 'Take Attendance' to begin.")
            self.attendance_button.configure(state='normal')

    def encode_faces(self, folder):
        list_people_encoding=[]

        for filename in os.listdir(folder):
            known_image=fr.load_image_file(f'{folder}{filename}')
            known_encoding=fr.face_encodings(known_image)[0]

            list_people_encoding.append((known_encoding,filename))

        return list_people_encoding

    def take_attendance(self):
        self.label.configure(text="Processing image...")
        self.master.update()

        for person in self.encode_faces('people/'):
            encoded_faces = person[0]
            filename = person[1]

            is_target_face = fr.compare_faces(encoded_faces, self.target_encoding, tolerance=0.55)
            presence = "Present" if True in is_target_face else "Absent"
            print(f'{presence} {filename}')

            if is_target_face:
                self.write_attendance(filename, presence)

        self.label.configure(text="Attendance taken and saved to file.")
        self.show_attendance_button.configure(state='normal')

    def write_attendance(self, name, attendance):
        wb = openpyxl.load_workbook("face_recognition_results.xlsx")
        sheet = wb.active

        row = sheet.max_row + 1
        sheet.cell(row, 1, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        sheet.cell(row, 2, name)
        sheet.cell(row, 3, attendance)

        wb.save("face_recognition_results.xlsx")

    def show_attendance(self):
        wb = openpyxl.load_workbook("face_recognition_results.xlsx")
        sheet = wb.active

        attendance_text = ""
        for row in sheet.iter_rows(values_only=True):
            attendance_text += f"{row[0]} - {row[1]} - {row[2]}\n"

        messagebox.showinfo("Attendance", attendance_text)

    def run_gui(self):
        self.master.mainloop()

if __name__ == "__main__":
    root = Tk()
    gui = FaceRecognitionGUI(root)
    root.mainloop()

